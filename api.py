#!/usr/bin/env python3
"""API routes for Renovation Expense Tracker."""
import os
import csv
import json
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, session
from functools import wraps
from models import db, Expense, KanbanCategory, Budget

api = Blueprint("api", __name__)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "未登录"}), 401
        return f(*args, **kwargs)
    return decorated


@api.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "已注销"})


# ==================== 支出 CRUD ====================

@api.route("/api/expenses", methods=["GET"])
@login_required
def get_expenses():
    category = request.args.get("category")
    status = request.args.get("status")
    sort_by = request.args.get("sort_by", "date")
    
    query = Expense.query
    if category:
        query = query.filter_by(category=category)
    if status:
        query = query.filter_by(status=status)
    
    if sort_by == "amount":
        query = query.order_by(Expense.amount.desc())
    else:
        query = query.order_by(db.desc(Expense.date))
    
    expenses = query.all()
    return jsonify({"expenses": [e.to_dict() for e in expenses], "count": len(expenses)})


@api.route("/api/expenses", methods=["POST"])
@login_required
def create_expense():
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供 JSON 数据"}), 400
    
    required = ["title", "category", "amount", "date"]
    for field in required:
        if field not in data:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400
    
    expense = Expense(
        title=data["title"],
        category=data["category"],
        amount=float(data["amount"]),
        date=data["date"],
        description=data.get("description", ""),
        status=data.get("status", "进行中"),
        priority=data.get("priority", "中"),
        area=data.get("area", "全屋"),
    )
    db.session.add(expense)
    db.session.commit()
    update_category_spent(expense.category, expense.amount, action="add")
    return jsonify({"expense": expense.to_dict()}), 201


@api.route("/api/expenses/<int:expense_id>", methods=["PUT"])
@login_required
def update_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供 JSON 数据"}), 400
    
    old_category = expense.category
    old_amount = expense.amount
    
    for field in ["title", "category", "amount", "date", "description", "status", "priority", "area"]:
        if field in data:
            setattr(expense, field, data[field])
    
    expense.updated_at = datetime.utcnow()
    db.session.commit()
    
    if old_category != expense.category or old_amount != expense.amount:
        update_category_spent(old_category, old_amount, action="subtract")
        update_category_spent(expense.category, expense.amount, action="add")
    
    return jsonify({"expense": expense.to_dict()})


@api.route("/api/expenses/<int:expense_id>", methods=["DELETE"])
@login_required
def delete_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    update_category_spent(expense.category, expense.amount, action="subtract")
    db.session.delete(expense)
    db.session.commit()
    return jsonify({"message": "已删除"})


# ==================== 看板分类 ====================

@api.route("/api/categories", methods=["GET"])
@login_required
def get_categories():
    categories = KanbanCategory.query.order_by(KanbanCategory.sort_order).all()
    return jsonify({"categories": [c.to_dict() for c in categories]})


@api.route("/api/categories", methods=["POST"])
@login_required
def create_category():
    data = request.get_json()
    if not data or "name" not in data:
        return jsonify({"error": "缺少分类名称"}), 400
    if KanbanCategory.query.filter_by(name=data["name"]).first():
        return jsonify({"error": "分类已存在"}), 400
    cat = KanbanCategory(
        name=data["name"],
        color=data.get("color", "#3498db"),
        sort_order=data.get("sort_order", 0),
    )
    db.session.add(cat)
    db.session.commit()
    return jsonify({"category": cat.to_dict()}), 201


@api.route("/api/categories/<int:cat_id>", methods=["PUT"])
@login_required
def update_category(cat_id):
    cat = KanbanCategory.query.get_or_404(cat_id)
    data = request.get_json()
    for field in ["name", "color", "sort_order"]:
        if field in data:
            setattr(cat, field, data[field])
    db.session.commit()
    return jsonify({"category": cat.to_dict()})


@api.route("/api/categories/<int:cat_id>", methods=["DELETE"])
@login_required
def delete_category(cat_id):
    cat = KanbanCategory.query.get_or_404(cat_id)
    if Expense.query.filter_by(category=cat.name).first():
        return jsonify({"error": "该分类下有支出记录，无法删除"}), 400
    db.session.delete(cat)
    db.session.commit()
    return jsonify({"message": "已删除"})


# ==================== 看板拖拽更新 ====================

@api.route("/api/board/move", methods=["POST"])
@login_required
def move_card():
    data = request.get_json()
    if not data or "moves" not in data:
        return jsonify({"error": "缺少 moves 数据"}), 400
    
    results = []
    for move in data["moves"]:
        expense = Expense.query.get(move.get("id"))
        if not expense:
            results.append({"id": move.get("id"), "status": "not_found"})
            continue
        
        old_category = expense.category
        if move.get("category") and move["category"] != old_category:
            amount = expense.amount
            update_category_spent(old_category, amount, action="subtract")
            expense.category = move["category"]
            update_category_spent(move["category"], amount, action="add")
        
        if move.get("status"):
            expense.status = move["status"]
        if move.get("priority"):
            expense.priority = move["priority"]
        
        expense.updated_at = datetime.utcnow()
        results.append({"id": expense.id, "status": "ok", "expense": expense.to_dict()})
    
    db.session.commit()
    return jsonify({"results": results})


# ==================== 预算统计 ====================

@api.route("/api/stats/summary", methods=["GET"])
@login_required
def get_summary():
    all_expenses = Expense.query.all()
    total_amount = sum(e.amount for e in all_expenses)
    total_budget = sum(b.total_budget for b in Budget.query.all())
    
    category_stats = {}
    for e in all_expenses:
        if e.category not in category_stats:
            category_stats[e.category] = {"total": 0, "count": 0}
        category_stats[e.category]["total"] += e.amount
        category_stats[e.category]["count"] += 1
    
    status_stats = {}
    for e in all_expenses:
        status_stats[e.status] = status_stats.get(e.status, 0) + 1
    
    return jsonify({
        "total_amount": round(total_amount, 2),
        "total_budget": total_budget,
        "budget_remaining": round(total_budget - total_amount, 2),
        "budget_usage": round(total_amount / total_budget * 100, 1) if total_budget > 0 else 0,
        "total_items": len(all_expenses),
        "category_stats": {k: {"total": round(v["total"], 2), "count": v["count"]} for k, v in category_stats.items()},
        "status_stats": status_stats,
    })


@api.route("/api/stats/categories", methods=["GET"])
@login_required
def get_category_stats():
    all_expenses = Expense.query.all()
    
    # Calculate actual spent from expenses table for each category
    actual_spent = {}
    for e in all_expenses:
        if e.category not in actual_spent:
            actual_spent[e.category] = 0
        actual_spent[e.category] += e.amount
    
    budgets = Budget.query.all()
    result_budgets = []
    for b in budgets:
        actual = actual_spent.get(b.category, 0)
        result_budgets.append({
            "id": b.id,
            "category": b.category,
            "total_budget": b.total_budget,
            "spent": round(actual, 2),
            "remaining": round(b.total_budget - actual, 2),
            "usage_rate": round(min(actual / b.total_budget * 100, 200), 1) if b.total_budget > 0 else 0,
        })
    
    return jsonify({"budgets": result_budgets})


@api.route("/api/stats/categories", methods=["POST"])
@login_required
def update_category_budget():
    data = request.get_json()
    if not data or "category" not in data or "total_budget" not in data:
        return jsonify({"error": "缺少 category 或 total_budget"}), 400
    
    budget = Budget.query.filter_by(category=data["category"]).first()
    if not budget:
        budget = Budget(category=data["category"], total_budget=float(data["total_budget"]))
        db.session.add(budget)
    else:
        budget.total_budget = float(data["total_budget"])
    
    db.session.commit()
    return jsonify({"budget": budget.to_dict()})


# ==================== 预算编辑 ====================

@api.route("/api/budget/update", methods=["PUT"])
@login_required
def update_budget():
    data = request.get_json()
    if not data or "category" not in data or "total_budget" not in data:
        return jsonify({"error": "缺少 category 或 total_budget"}), 400
    
    budget = Budget.query.filter_by(category=data["category"]).first()
    if not budget:
        budget = Budget(category=data["category"], total_budget=float(data["total_budget"]))
        db.session.add(budget)
    else:
        budget.total_budget = float(data["total_budget"])
    
    db.session.commit()
    return jsonify({"success": True, "budget": budget.to_dict()})


def update_category_spent(category, amount, action="add"):
    budget = Budget.query.filter_by(category=category).first()
    if not budget:
        if action == "add":
            budget = Budget(category=category, total_budget=0, spent=amount)
            db.session.add(budget)
    else:
        if action == "add":
            budget.spent += amount
        elif action == "subtract":
            budget.spent = max(0, budget.spent - amount)
    db.session.commit()


# ==================== 数据导入 ====================

@api.route("/api/import/csv", methods=["POST"])
@login_required
def import_csv():
    if "file" not in request.files:
        return jsonify({"error": "请上传文件"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "请选择文件"}), 400
    
    if not file.filename.endswith((".csv", ".xlsx", ".xls")):
        return jsonify({"error": "请上传 CSV 或 Excel 文件"}), 400
    
    try:
        content = file.read()
        rows = []
        
        if file.filename.endswith(".csv"):
            for encoding in ["utf-8", "gbk", "gb2312", "gb18030"]:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                text = content.decode("utf-8", errors="ignore")
            
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                normalized = normalize_row(row)
                if "title" in normalized and "amount" in normalized:
                    rows.append(normalized)
        
        elif file.filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "Excel 格式需要安装 openpyxl 库"}), 500
            
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active
            headers = [str((cell.value or "")).strip() for cell in sheet[1]]
            
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                    if cell_value is not None:
                        row_data[header] = str(cell_value).strip()
                
                normalized = normalize_row(row_data)
                if "title" in normalized and "amount" in normalized:
                    rows.append(normalized)
        
        if not rows:
            return jsonify({"error": "未解析到有效数据，请检查文件内容"}), 400
        
        for row in rows:
            expense = Expense(
                title=row["title"],
                category=row["category"],
                amount=float(row["amount"]),
                date=row["date"],
                description=row.get("description", ""),
                status=row.get("status", "进行中"),
                priority=row.get("priority", "中"),
                area=row.get("area", "全屋"),
            )
            db.session.add(expense)
        
        db.session.commit()
        
        for row in rows:
            update_category_spent(row["category"], float(row["amount"]), "add")
        
        return jsonify({"message": f"成功导入 {len(rows)} 条记录", "imported": len(rows)})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"导入失败: {str(e)}"}), 500


def parse_chinese_date(date_str):
    """解析中文日期格式: 2026年4月19日 08:18 → 2026-04-19"""
    if not date_str:
        return None
    date_str = str(date_str).strip()
    # 尝试 "2026年4月19日" 或 "2026年4月19日 08:18" 格式
    import re
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # 尝试标准日期格式
    try:
        return datetime.strptime(date_str[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return date_str[:10]


def normalize_row(row):
    """标准化字段名"""
    # Strip BOM from all keys
    row = {k.lstrip('\ufeff') if isinstance(k, str) else k: v for k, v in row.items()}
    
    normalized = {}
    # 特殊映射：费用类别 → 分类（带映射）
    raw_category = None
    if "费用类别" in row and (row["费用类别"] or "").strip():
        raw_category = (row["费用类别"] or "").strip()
    if "区域" in row and (row["区域"] or "").strip() and "category" not in normalized:
        normalized["area"] = (row["区域"] or "").strip()
    
    # 特殊映射：费用 → 金额
    if "费用" in row and (row["费用"] or "").strip():
        v = (row["费用"] or "").strip()
        try:
            normalized["amount"] = float(v.replace(",", "").replace("¥", "").replace("￥", ""))
        except (ValueError, TypeError):
            normalized["amount"] = 0
    
    # 特殊映射：创建时间 → 日期（优先于"日期"字段）
    if "创建时间" in row and (row["创建时间"] or "").strip():
        parsed = parse_chinese_date(row["创建时间"])
        if parsed:
            normalized["date"] = parsed
    
    # 映射费用类别到系统分类
    category_mapping = {
        # 人工费 → 人工费用
        "人工费": "人工费用",
        # 材料费 → 根据装修阶段进一步映射
        # 大电器 → 电器采购
        "大电器": "电器采购",
        # 工具费 → 其他支出
        "工具费": "其他支出",
        # 租赁费 → 其他支出
        "租赁费": "其他支出",
    }
    
    if raw_category and raw_category in category_mapping:
        normalized["category"] = category_mapping[raw_category]
    elif raw_category:
        normalized["category"] = raw_category
    
    for k, v in row.items():
        k = (k or "").strip()
        v = (v or "").strip()
        if not v:
            continue
        
        if k in ["名称", "项目", "title", "name"]:
            normalized["title"] = v
        elif k in ["分类", "category", "type"]:
            if "category" not in normalized:
                normalized["category"] = v
        elif k in ["金额", "价格", "amount", "price", "cost"]:
            if "amount" not in normalized:
                try:
                    normalized["amount"] = float(v.replace(",", "").replace("¥", "").replace("￥", ""))
                except (ValueError, TypeError):
                    normalized["amount"] = 0
        elif k in ["日期", "date", "时间"]:
            if "date" not in normalized:
                parsed = parse_chinese_date(v)
                if parsed:
                    normalized["date"] = parsed
        elif k in ["备注", "说明", "description", "notes"]:
            normalized["description"] = v
        elif k in ["状态", "status"]:
            normalized["status"] = v
        elif k in ["优先级", "priority", "紧急程度"]:
            normalized["priority"] = v
    
    normalized.setdefault("date", datetime.now().strftime("%Y-%m-%d"))
    normalized.setdefault("category", "其他支出")
    normalized.setdefault("status", "进行中")
    normalized.setdefault("priority", "中")
    normalized.setdefault("area", "全屋")
    return normalized


# ==================== 示例数据 ====================

@api.route("/api/sample-data", methods=["POST"])
@login_required
def load_sample_data():
    sample_expenses = [
        ("拆除工程", "泥瓦工程", 5000, "2025-01-15", "原有墙面拆除、地面拆除"),
        ("水电改造材料", "水电改造", 12000, "2025-01-20", "电线、水管、开关面板等"),
        ("防水工程", "泥瓦工程", 3000, "2025-01-25", "卫生间、厨房防水处理"),
        ("瓷砖（客厅）", "瓷砖铺贴", 8500, "2025-02-01", "800x800 抛光砖约80平米"),
        ("瓷砖（厨卫）", "瓷砖铺贴", 4500, "2025-02-10", "300x300 防滑砖"),
        ("全屋定制衣柜", "木工工程", 18000, "2025-02-20", "主卧+次卧+书房，E0级板材"),
        ("橱柜", "木工工程", 12000, "2025-02-25", "石英石台面+多层实木板"),
        ("乳胶漆", "油漆工程", 6000, "2025-03-01", "多乐士森呼吸系列"),
        ("全屋吊顶", "木工工程", 15000, "2025-03-05", "客厅+餐厅+走廊石膏板吊顶"),
        ("室内门×4扇", "泥瓦工程", 8000, "2025-03-10", "实木复合门含五金"),
        ("卫生间洁具", "电器采购", 6500, "2025-03-15", "马桶+花洒+浴室柜+五金"),
        ("中央空调", "电器采购", 28000, "2025-03-20", "一拖四，格力风管机"),
        ("地暖", "水电改造", 16000, "2025-03-25", "壁挂炉+地暖管+分水器"),
        ("厨房烟机灶具", "电器采购", 5500, "2025-04-01", "方太烟机灶套装"),
        ("客厅沙发", "家具购置", 8000, "2025-04-10", "真皮三人位沙发"),
        ("餐桌椅", "家具购置", 4500, "2025-04-15", "岩板餐桌+6椅"),
        ("床+床垫×2", "家具购置", 12000, "2025-04-20", "主卧+次卧实木床"),
        ("窗帘+纱帘", "软装搭配", 4000, "2025-04-25", "全屋遮光帘+纱帘"),
        ("灯具", "电器采购", 6000, "2025-05-01", "客厅主灯+卧室灯+筒灯"),
        ("家电（电视/冰箱/洗衣机）", "电器采购", 22000, "2025-05-10", "55寸电视+对开门冰箱+滚筒洗衣机"),
        ("开荒保洁", "其他支出", 1500, "2025-05-15", "全屋深度清洁"),
        ("甲醛治理", "其他支出", 2000, "2025-05-18", "专业机构全屋治理"),
        ("装饰画+摆件", "软装搭配", 1500, "2025-05-20", "客厅+餐厅装饰画"),
        ("窗帘轨道+五金", "其他支出", 800, "2025-05-22", "全屋轨道安装"),
        ("垃圾清运", "其他支出", 1200, "2025-05-25", "装修垃圾运出小区"),
    ]
    
    for title, category, amount, date, desc in sample_expenses:
        if not Expense.query.filter_by(title=title).first():
            expense = Expense(
                title=title, category=category, amount=amount,
                date=date, description=desc, status="已完成", priority="中",
            )
            db.session.add(expense)
    
    default_budgets = {
        "水电改造": 30000, "泥瓦工程": 40000, "木工工程": 50000,
        "油漆工程": 10000, "瓷砖铺贴": 15000, "家具购置": 30000,
        "电器采购": 40000, "软装搭配": 6000, "其他支出": 6000,
    }
    
    for cat_name, budget in default_budgets.items():
        budget_obj = Budget.query.filter_by(category=cat_name).first()
        if not budget_obj:
            budget_obj = Budget(category=cat_name, total_budget=budget)
            db.session.add(budget_obj)
    
    db.session.commit()
    return jsonify({"message": f"已加载 {len(sample_expenses)} 条示例数据"})


# ==================== 图表数据 ====================

@api.route("/api/charts/by-area", methods=["GET"])
@login_required
def chart_by_area():
    """按区域汇总支出"""
    from sqlalchemy import func
    result = db.session.query(
        Expense.area,
        func.sum(Expense.amount).label("total")
    ).group_by(Expense.area).all()
    return jsonify({
        "data": [{"name": r.area or "未分类", "value": float(r.total)} for r in result]
    })


@api.route("/api/charts/by-category", methods=["GET"])
@login_required
def chart_by_category():
    """按分类汇总支出"""
    from sqlalchemy import func
    result = db.session.query(
        Expense.category,
        func.sum(Expense.amount).label("total")
    ).group_by(Expense.category).all()
    return jsonify({
        "data": [{"name": r.category, "value": float(r.total)} for r in result]
    })


@api.route("/api/charts/by-type", methods=["GET"])
@login_required
def chart_by_type():
    """人工费用 vs 其他分类汇总"""
    from sqlalchemy import func
    labor = db.session.query(func.sum(Expense.amount)).filter(
        Expense.category == "人工费用"
    ).scalar() or 0
    other = db.session.query(func.sum(Expense.amount)).filter(
        Expense.category != "人工费用"
    ).scalar() or 0
    return jsonify({
        "data": [
            {"name": "人工费用", "value": float(labor)},
            {"name": "其他类别", "value": float(other)},
        ]
    })
